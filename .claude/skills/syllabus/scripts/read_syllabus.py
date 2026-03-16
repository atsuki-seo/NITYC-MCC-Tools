#!/usr/bin/env python3
"""シラバスエクセルファイルの読み取りスクリプト

エクセルファイルからシラバス情報を抽出し、JSON形式で標準出力に出力する。
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxlがインストールされていません。pip3 install openpyxl を実行してください。"}))
    sys.exit(1)


def read_syllabus(filepath: str) -> dict:
    path = Path(filepath)
    if not path.exists():
        return {"error": f"ファイルが見つかりません: {filepath}"}

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {"error": f"ファイルの読み込みに失敗しました: {e}"}

    expected_sheets = [
        "科目情報・到達目標",
        "ルーブリック",
        "教育方法等・自由使用枠",
        "授業計画",
        "評価割合",
        "授業の属性・履修上の区分",
    ]
    missing = [s for s in expected_sheets if s not in wb.sheetnames]
    if missing:
        return {"error": f"以下のシートが見つかりません: {', '.join(missing)}", "found_sheets": wb.sheetnames}

    result = {"sheets": wb.sheetnames}

    # --- 科目情報・到達目標 ---
    ws = wb["科目情報・到達目標"]
    result["basic_info"] = {
        "教科名": ws["B1"].value,
        "科目番号": ws["B4"].value,
        "科目区分": ws["D4"].value,
        "授業形態": ws["B5"].value,
        "単位": ws["D5"].value,
        "開設学科": ws["B6"].value,
        "対象学年": ws["D6"].value,
        "開講期": ws["B7"].value,
        "週時間数": ws["D7"].value,
        "教科書_教材": ws["B8"].value,
        "担当教員": ws["B9"].value,
    }
    result["到達目標"] = ws["A13"].value
    result["学科の到達目標項目との関係"] = ws["A16"].value

    # --- ルーブリック ---
    ws = wb["ルーブリック"]
    rubrics = []
    for row in range(3, ws.max_row + 1):
        aspect = ws.cell(row=row, column=1).value
        if not aspect:
            break
        rubrics.append({
            "評価観点": aspect,
            "理想的": ws.cell(row=row, column=2).value,
            "標準的": ws.cell(row=row, column=3).value,
            "未到達": ws.cell(row=row, column=4).value,
        })
    result["ルーブリック"] = rubrics

    # --- 教育方法等・自由使用枠 ---
    ws = wb["教育方法等・自由使用枠"]
    result["教育方法等"] = {
        "概要": ws["B3"].value,
        "授業の進め方": ws["B4"].value,
        "注意点": ws["B5"].value,
    }

    # --- 授業計画 ---
    ws = wb["授業計画"]
    plan = []
    # 前期: rows 3-18, 後期: rows 19-34
    for start_row, semester in [(3, "前期"), (19, "後期")]:
        for i in range(16):
            row = start_row + i
            week_label = ws.cell(row=row, column=3).value
            content = ws.cell(row=row, column=4).value
            goal = ws.cell(row=row, column=5).value
            if week_label or content or goal:
                plan.append({
                    "学期": semester,
                    "週": week_label,
                    "授業内容": content,
                    "到達目標": goal,
                })
    result["授業計画"] = plan

    # --- 評価割合 ---
    ws = wb["評価割合"]
    methods = []
    for col in range(2, 9):  # B-H
        val = ws.cell(row=1, column=col).value
        methods.append(val)

    perspectives = []
    for row in range(2, 6):  # rows 2-5
        name = ws.cell(row=row, column=1).value
        values = {}
        for ci, method in enumerate(methods):
            val = ws.cell(row=row, column=ci + 2).value
            values[method or f"col{ci+2}"] = val
        perspectives.append({"観点": name, "配分": values})
    result["評価割合"] = {"評価方法": methods, "観点別配分": perspectives}

    # --- 授業の属性・履修上の区分 ---
    ws = wb["授業の属性・履修上の区分"]
    result["授業の属性"] = {
        "アクティブラーニング": ws["B2"].value,
        "ICT利用": ws["B3"].value,
        "遠隔授業対応": ws["B4"].value,
    }

    return result


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "使用法: python3 read_syllabus.py <エクセルファイルパス>"}))
        sys.exit(1)

    result = read_syllabus(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))

    if "error" in result:
        sys.exit(1)


if __name__ == "__main__":
    main()
