#!/usr/bin/env python3
"""シラバスエクセルファイルへの書き込みスクリプト

JSON形式のデータを受け取り、エクセルファイルの該当セルに書き込む。
既存の書式（フォント、罫線、セル結合）は保持する。
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxlがインストールされていません。pip3 install openpyxl を実行してください。"}))
    sys.exit(1)


def write_syllabus(filepath: str, data: dict) -> dict:
    path = Path(filepath)
    if not path.exists():
        return {"error": f"ファイルが見つかりません: {filepath}"}

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        return {"error": f"ファイルの読み込みに失敗しました: {e}"}

    written = []

    # --- 教育方法等・自由使用枠 ---
    if "教育方法等" in data:
        ws = wb["教育方法等・自由使用枠"]
        info = data["教育方法等"]
        if "概要" in info and info["概要"] is not None:
            ws["B3"] = info["概要"]
            written.append("教育方法等.概要")
        if "注意点" in info and info["注意点"] is not None:
            ws["B5"] = info["注意点"]
            written.append("教育方法等.注意点")

    # --- 授業計画 ---
    if "授業計画" in data:
        ws = wb["授業計画"]
        plan = data["授業計画"]

        # 開講期を判定して書き込み開始行を決定
        semester = plan.get("学期", "後期")
        start_row = 3 if semester == "前期" else 19

        for week in plan.get("週データ", []):
            week_num = week.get("週番号", 0)
            if week_num < 1:
                continue
            row = start_row + week_num - 1
            if "授業内容" in week:
                ws.cell(row=row, column=4, value=week["授業内容"])
            if "到達目標" in week:
                ws.cell(row=row, column=5, value=week["到達目標"])
        written.append(f"授業計画（{len(plan.get('週データ', []))}週分）")

    # --- 評価割合 ---
    if "評価割合" in data:
        ws = wb["評価割合"]
        eval_data = data["評価割合"]

        for perspective in eval_data.get("観点別配分", []):
            row_map = {
                "知識の基本的な理解": 2,
                "汎用的技能": 3,
                "主体的・継続的な学習意欲": 4,
                "態度・指向性": 5,
            }
            row = row_map.get(perspective.get("観点"))
            if row is None:
                continue

            method_col_map = {
                "定期試験": 2, "小テスト": 3, "レポート": 4,
                "口頭発表": 5, "成果物実技": 6, "ポートフォリオ": 7, "その他": 8,
            }
            for method, value in perspective.get("配分", {}).items():
                col = method_col_map.get(method)
                if col is not None:
                    ws.cell(row=row, column=col, value=value)
        written.append("評価割合")

    # --- 授業の属性 ---
    if "授業の属性" in data:
        ws = wb["授業の属性・履修上の区分"]
        attrs = data["授業の属性"]
        attr_map = {"アクティブラーニング": "B2", "ICT利用": "B3", "遠隔授業対応": "B4"}
        for key, cell in attr_map.items():
            if key in attrs:
                ws[cell] = attrs[key]
        written.append("授業の属性")

    try:
        wb.save(filepath)
    except Exception as e:
        return {"error": f"ファイルの保存に失敗しました: {e}"}

    return {"success": True, "written": written}


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "使用法: python3 write_syllabus.py <エクセルファイルパス> <JSONデータ>"}))
        sys.exit(1)

    filepath = sys.argv[1]
    try:
        data = json.loads(sys.argv[2])
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"JSONの解析に失敗しました: {e}"}))
        sys.exit(1)

    result = write_syllabus(filepath, data)
    print(json.dumps(result, ensure_ascii=False, indent=2))

    if "error" in result:
        sys.exit(1)


if __name__ == "__main__":
    main()
